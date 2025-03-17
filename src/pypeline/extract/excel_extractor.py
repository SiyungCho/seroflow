"""
"""
import xlrd
from openpyxl import load_workbook
import pandas as pd
from ..extract.file_extractor import FileExtractor, MultiFileExtractor

class ExcelExtractor(FileExtractor):
    """
    """

    def __init__(self,
                 source,
                 step_name="ExcelExtractor",
                 chunk_size=None,
                 on_error=None,
                 **kwargs):
        """
        """
        super().__init__(source=source,
                         step_name=step_name,
                         func = self.func if chunk_size is None else self.chunk_func,
                         chunk_size=chunk_size,
                         on_error=on_error,
                         **kwargs)

    def func(self, context):
        """
        """
        context.add_dataframe(self.file_name, self.__read_excel(self.file_path, self.kwargs))
        return context

    def chunk_func(self, context, chunk_coordinates):
        """
        """
        context.add_dataframe(self.file_name, self.__read_excel_chunk(self.file_path, chunk_coordinates, self.kwargs))
        return context

    def __read_excel(self, file, kwargs):
        """
        """
        if file.endswith('.xls'):
            return pd.read_excel(file, engine='xlrd', **kwargs)
        if file.endswith('.xlsx'):
            return pd.read_excel(file, engine='openpyxl', **kwargs)
        raise ValueError(f"Unsupported file format: {file}")

    def __read_excel_chunk(self, file, chunk_coordinates, kwargs):
        """
        """
        start_idx, stop_idx = chunk_coordinates
        if start_idx is None:
            return pd.DataFrame()

        nrows = stop_idx - start_idx

        if file.endswith('.xls'):
            engine = 'xlrd'
        elif file.endswith('.xlsx'):
            engine = 'openpyxl'
        else:
            raise ValueError(f"Unsupported file format: {file}")

        nrows = stop_idx - start_idx
        return pd.read_excel(file, skiprows=start_idx, nrows=nrows, engine=engine, **kwargs)

    def get_max_row_count(self):
        """
        """
        max_rows = 0
        if self.file_path.endswith('.xlsx'):
            wb = load_workbook(filename=self.file_path, read_only=True)
            ws = wb.active  # use the first (active) sheet
            rows_count = ws.max_row
            wb.close()
        elif self.file_path.endswith('.xls'):
            wb = xlrd.open_workbook(self.file_path, on_demand=True)
            ws = wb.sheet_by_index(0)
            rows_count = ws.nrows
            wb.release_resources()
        else:
            raise ValueError(f"Unsupported file format: {self.file_path}")

        max_rows = max(max_rows, rows_count)
        return max_rows

class MultiExcelExtractor(MultiFileExtractor):
    """
    """
    def __init__(self,
                 source,
                 chunk_size=None,
                 on_error=None,
                 **kwargs):
        """
        """
        super().__init__(source=source,
                         step_name="MultiExcelExtractor",
                         type=ExcelExtractor,
                         extension_type='excel',
                         chunk_size=chunk_size,
                         on_error=on_error,
                         **kwargs)

    def get_max_row_count(self):
        """
        """
        max_rows = 0
        for file in self.file_paths:
            if file.endswith('.xlsx'):
                wb = load_workbook(filename=file, read_only=True)
                ws = wb.active  # use the first (active) sheet
                rows_count = ws.max_row
                wb.close()
            elif file.endswith('.xls'):
                wb = xlrd.open_workbook(file, on_demand=True)
                ws = wb.sheet_by_index(0)
                rows_count = ws.nrows
                wb.release_resources()
            else:
                raise ValueError(f"Unsupported file format: {file}")

            max_rows = max(max_rows, rows_count)

        return max_rows
